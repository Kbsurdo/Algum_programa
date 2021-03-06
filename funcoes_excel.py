from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.styles.borders import Side
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Alignment, Font
import random


def criar_pasta_trabalho():
    wb = Workbook()
    return wb


def abrir_pasta_trabalho(nome_pasta_trabalho):
    wb = load_workbook(nome_pasta_trabalho+'.xlsx')
    return wb


def copiar_exemplo():
    wb = abrir_pasta_trabalho('EXEMPLO_MAO')
    return wb


def planilha_ativa(wb, nome_planilha=None):
    if nome_planilha is None:
        planilha = wb.active
        return planilha
    else:
        planilha = wb[nome_planilha]
        return planilha


def criar_planilha(wb, nome_planilha):
    wb.create_sheet(nome_planilha)


def mudar_nome_planilha(wb, novo_nome_planilha, nome_planilha=None):
    planilha_ativa(wb, nome_planilha).title = novo_nome_planilha


def get_dados(beg_row, end_row, beg_col, end_col, planilha):
    for row in range(beg_row, end_row):
        for col in range(beg_col, end_col):
            char = get_column_letter(col)
            print(planilha[char + str(row)].value, end=',')
        print('\n')


def criar_cabecario_mao(wb, nome_pasta_trabalho, rev, nome_planilha=None):
    ws = planilha_ativa(wb, nome_planilha)
    borda_fina = Side(border_style='thin', color='000000')
    borda_grossa = Side(border_style='medium', color='000000')

    ws.column_dimensions['A'].width = 12+0.71
    ws.column_dimensions['B'].width = 9+0.71
    ws.column_dimensions['C'].width = 10.86+0.71
    ws.column_dimensions['D'].width = 15+0.71
    ws.column_dimensions['E'].width = 12+0.71
    ws.column_dimensions['F'].width = 12+0.71
    ws.column_dimensions['G'].width = 12+0.71
    ws.column_dimensions['H'].width = 12+0.71
    ws.column_dimensions['I'].width = 12+0.71
    ws.column_dimensions['J'].width = 8+0.71

    ws.merge_cells('A1:C2')
    ws.merge_cells('D1:I2')
    ws.merge_cells('J1:J2')
    ws.merge_cells('A4:A5')
    ws.merge_cells('B4:B5')
    ws.merge_cells('C4:C5')
    ws.merge_cells('D4:D5')
    ws.merge_cells('E4:I5')
    ws.merge_cells('J4:J5')

    for row in range(1, 3):
        for col in range(1, 11):
            char = get_column_letter(col)
            ws[f'{char}{row}'].font = Font(name='Arial', size=11, bold=True, color='000000')
            if col == 1:
                ws[f'{char}{str(row)}'].border = Border(top=borda_grossa, left=borda_grossa, bottom=borda_grossa)
                ws[f'{char}{str(row)}'].alignment = Alignment(horizontal='left', vertical='center')
            elif col == 10:
                ws[f'{char}{str(row)}'].border = Border(top=borda_grossa, right=borda_grossa, bottom=borda_grossa)
                ws[f'{char}{str(row)}'].alignment = Alignment(horizontal='center', vertical='center')
            else:
                ws[f'{char}{str(row)}'].border = Border(top=borda_grossa, bottom=borda_grossa)
                ws[f'{char}{str(row)}'].alignment = Alignment(horizontal='center', vertical='center')

    for row in range(4, 6):
        for col in range(1, 11):
            char = get_column_letter(col)
            ws[f'{char}{str(row)}'].border = Border(top=borda_fina, left=borda_fina, right=borda_fina, bottom=borda_fina)
            ws[f'{char}{str(row)}'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
            ws[f'{char}{row}'].font = Font(name='Arial', size=11, bold=True, color='000000')

    nome_lista_cabo = nome_pasta_trabalho
    inicio = nome_lista_cabo.index('MAO')
    fim = nome_lista_cabo.index('LC') + 2
    aux1 = nome_lista_cabo[inicio:fim]
    ws['A1'] = 'AMAZONAS ENERGIA - DTE'
    ws['D1'] = 'LISTA DE CABOS' + (f'   {aux1}')
    ws['J1'] = (f'REV. {rev}')

    ws['A4'] = 'N CABO'
    ws['B4'] = 'COMP.\n(m)'
    ws['C4'] = 'FORM.'
    ws['D4'] = 'ORIGEM\nDESTINO'
    ws['E4'] = 'PERCURSO'
    ws['J4'] = 'REV'


def criar_cabo_mao(wb, nome_planilha, linha, quant_dados):
    ws = planilha_ativa(wb, nome_planilha)
    borda_fina = Side(border_style='thin', color='000000')
    for linha_indice in range(int(linha), int(linha)+(int(quant_dados)*2), 2):
        for col in range(1, 11):
            char = get_column_letter(col)
            if col < 5 or col == 10:
                ws.merge_cells(
                    f'{char}{linha_indice}:{char}{str(int(linha_indice)+1)}')
                for aux1 in range(0, 2):
                    ws[f'{char}{str(int(linha_indice)+aux1)}'].border = Border(top=borda_fina, left=borda_fina, right=borda_fina, bottom=borda_fina)
                    ws[f'{char}{str(int(linha_indice)+aux1)}'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                    ws[f'{char}{str(int(linha_indice)+aux1)}'].font = Font(name='Arial', size=11, color='000000')
            else:
                for aux2 in range(0, 2):
                    ws[f'{char}{str(int(linha_indice)+aux2)}'].border = Border(top=borda_fina, bottom=borda_fina)
                    ws[f'{char}{str(int(linha_indice)+aux2)}'].alignment = Alignment(horizontal='center', vertical='center', wrapText=True)
                    ws[f'{char}{str(int(linha_indice)+aux2)}'].font = Font(name='Arial', size=10, color='000000')


def preencher_cabo_mao(wb, nome_planilha, linha, tag_cabo, comprimento, formacao, origem_destino, percurso, rev):
    ws = planilha_ativa(wb, nome_planilha)
    for col in range(1, 11):
        char = get_column_letter(col)
        match char:
            case 'A':
                ws[f'{char}{str(linha)}'] = tag_cabo
            case 'B':
                ws[f'{char}{str(linha)}'] = comprimento
            case 'C':
                ws[f'{char}{str(linha)}'] = formacao
            case 'D':
                ws[f'{char}{str(linha)}'] = origem_destino
            case 'E':
                ws[f'{char}{str(linha)}'] = percurso[0]
                ws[f'{char}{str(linha+1)}'] = percurso[5]
            case 'F':
                ws[f'{char}{str(linha)}'] = percurso[1]
                ws[f'{char}{str(linha+1)}'] = percurso[6]
            case 'G':
                ws[f'{char}{str(linha)}'] = percurso[2]
                ws[f'{char}{str(linha+1)}'] = percurso[7]
            case 'H':
                ws[f'{char}{str(linha)}'] = percurso[3]
                ws[f'{char}{str(linha+1)}'] = percurso[8]
            case 'I':
                ws[f'{char}{str(linha)}'] = percurso[4]
                ws[f'{char}{str(linha+1)}'] = percurso[9]
            case 'J':
                ws[f'{char}{str(linha)}'] = rev
    return wb


if __name__ == '__main__':
    wb = copiar_exemplo()
    data = ['', '', '', '', '', '', '', '', '', '']
    criar_planilha(wb, 'Cabos')
    criar_cabecario_mao(wb, 'MAO-969-875010-LC', 'B', 'Cabos')
    criar_cabo_mao(wb, 'Cabos', '6', '250')
    for aux1 in range(6, 506, 2):
        for aux2 in range(10):
            data[aux2] = random.randint(1, 1000)
        aux3 = random.randint(2000, 3000)
        aux4 = random.randint(5, 300)
        wb = preencher_cabo_mao(wb, 'Cabos', aux1, f'1-CA1-{aux3}', f'{aux4}', '(2x4)', 'QSACC-02\nQSACA-01', data, 'A')
    wb.save('MAO-969-875010-LC-B - Lista de cabos de controle.xlsx')
